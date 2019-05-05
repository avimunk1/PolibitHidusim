#this utility gets an xls file export from Sumera to update the qullaInputObject

class sumeraPolicieshandler():
    """docstring for ."""

    def __init__(sumeraPoliciesObj):
        myid=id(sumeraPoliciesObj)
        sumeraPoliciesObj.sobjId=myid
        sumeraPoliciesObj.sumeraEgentId =sumeraEgentId
        sumeraPoliciesObj.sumeraAnefId=sumeraAnefId
        sumeraPoliciesObj.sumeraPolicyIdPreviuesYear=sumeraPolicyIdPreviuesYear
        sumeraPoliciesObj.sumeraPolicyIdThisYear=sumeraPolicyIdPreviuesYear
        sumeraPoliciesObj.sumeraValidetionCode=sumeraValidetionCode
        sumeraPoliciesObj.sumeraStatusInt=sumeraStatusInt
        sumeraPoliciesObj.sumeraStartDate=sumeraStartDate
        sumeraPoliciesObj.sumeraCustomerName=sumeraCustomerName
        sumeraPoliciesObj.sumeraCustomerId=sumeraCustomerId
        sumeraPoliciesObj.sumeraVehicleLicensePlate=sumeraVehicleLicensePlate
        sumeraPoliciesObj.sumeraDedectaboleMethod=sumeraDedectaboleMethod
        sumeraPoliciesObj.sumeraHoreaId=sumeraHoreaId
        sumeraPoliciesObj.sumeraNoOfPyments=sumeraNoOfPyments
        sumeraPoliciesObj.sumeraLastYearPremia=sumeraLastYearPremia
        sumeraPoliciesObj.sumeraThisYearPremia=sumeraThisYearPremia
        sumeraPoliciesObj.sumeraRenewErrorDescrption=sumeraRenewErrorDescrption
        #print(sumeraPoliciesObj.sobjId)


# all fields to be in use
sumeraEgentId ='' #מספר סוכן
sumeraAnefId='' #מספר ענף
sumeraPolicyIdPreviuesYear='' #פוליסה אשתקד
sumeraPolicyIdThisYear='' #פוליסה השנה
sumeraValidetionCode='' #בקרה
sumeraStatusInt='' #סטאטוס אחרון
sumeraStartDate='' #תחילת ביטוח
sumeraCustomerName='' #שם מבוטח
sumeraCustomerId='' #ת.ז
sumeraVehicleLicensePlate='' #מספר רישוי
sumeraDedectaboleMethod='' #שיטת גביה
sumeraHoreaId='' #מספר הוראה
sumeraNoOfPyments='' #מספר תשלומים
sumeraLastYearPremia='' #פרמיה אשתקד במזומן
sumeraThisYearPremia='' #פרמיה לגביה במזומן
sumeraRenewErrorDescrption='' #תאור שגיאה אי חידוש


#file hendling
iputfolder='/inputs'
inputFile="sumera_short-2.xlsx"
#outPutFile="fixedBafi.xlsx"

try:
    fileH=open('inputs/'+inputFile)
except:
    print('no such file'+inputFile)
    exit()

#prepering xls's for read and write
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
writeWB = Workbook()
rws = writeWB.active
from openpyxl import load_workbook
readWB = load_workbook('inputs/'+inputFile)
wws = readWB.active

# read the input xls, add valid mobilePhone
excelLine=1
excelColmn=0
mobilePhone=''
objList=dict()
sumerainfo=list()

for row in wws.values:
    sumeraEgentId=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraAnefId=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraPolicyIdPreviuesYear=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraPolicyIdThisYear=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraValidetionCode=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraStatusInt=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraStartDate=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraCustomerName=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraVehicleLicensePlate=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraDedectaboleMethod=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraHoreaId=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraNoOfPyments=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraLastYearPremia=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraThisYearPremia=row[excelColmn]
    excelColmn=excelColmn+1
    sumeraRenewErrorDescrption=row[excelColmn]
    excelLine=excelLine+1
    excelColmn=0
    uid=str(sumeraPolicyIdThisYear)+str(sumeraAnefId)

    myobject=sumeraPolicieshandler()
    sumerainfo[0]=uid
    sumerainfo[1]=myobject.sumeraThisYearPremia



    #mykey=myobject.sobjId
    #objList[mykey]=mykey
    #objList[uid]=mykey
a=sumerainfo.find('739135677919730')
#myrefrense=objList['739135677919730']
#a=myobject.sumeraEgentId
#a=get_objects()
#print(a)
