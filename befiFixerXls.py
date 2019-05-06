#this utility gets a csv export from Bafi and create an xls
# it handles the phones that are in one column in the original file

# all fields to be in use
status=''
customerName=''
customerID=''
phones=''
Address=''
email=''
policyNumber=''
insuranceType=''
insuranceCompny=''
anafID=''
anafName=''
policyStartDate=''
policyEndDate=''
prmia=''
veicleType=''
SubAnaf=''
carLicenessPlate=''
manifctureYear=''
vehicleModelId=''
vehicleModelDescription=''
allowedToDrive=''
protuction=''
propertyAddress=''
paymentType=''
numofPayments=''
egentName=''
egentId=''
ipdatedBy=''
crmRepName=''
salesRepName=''
comments=''
policyhandkerName=''
LocalsumeraLastYearPremia=''
policy=list()

def checkMobilePhone(phone): #check numbers and retren the ststus
    phone=str(phone)
    checkStatus='0'
    if not len(phone)==10:
        checkStatus='1'
    if not phone.startswith('05'):
        checkStatus='2'
    if not phone.isnumeric():
        checkStatus='3'
    return(checkStatus)


def phoneSplitter(phones): #handle the column "טלפונים" to get a valib mobilePhone
    phoneSplit=list()
    phoneStr=phones
    phone1=''

    chrcount=0
    mobilePhone=''
    status=''
    if len(phoneStr)<1:
        return('err')
        exit()
    if phoneStr.find(';')<0: #i.e a single phone in the string
        for char in phoneStr:
            chrcount=chrcount+1
            x=len(phoneStr)
            phoneStr=phoneStr.strip()
            if not char.isnumeric():
                continue
            else:
                phone1=phone1+char
            if x>0 and x==chrcount:
                phoneSplit.append(phone1)
                status=checkMobilePhone(phone1)
                if status=='0':
                    mobilePhone=phone1
                    return(mobilePhone)
                    exit()
                else:
                    phone1=''

    if phoneStr.find(';')>-1:# i.e moltipile phones in the string
        for char in phoneStr:
            phoneStr=phoneStr.strip()
            x=len(phoneStr)
            chrcount=chrcount+1

            if  char.isnumeric():
                phone1=phone1+char
            if char==';' or chrcount==x :
                phoneSplit.append(phone1)
                phone1=''

        phone04=''
        phone02=phoneSplit[0]
        status=checkMobilePhone(phone02)
        if status=='0':
            mobilePhone=phone02
            return(mobilePhone)
            exit()
        else:
            phone03=phoneSplit[1]
            status=checkMobilePhone(phone03)
            if status=='0':
                mobilePhone=phone03
                return(mobilePhone)
                exit()
            else:
                if len(phoneSplit)>2:
                    phone04=phoneSplit[2]
                    status=checkMobilePhone(phone04)
                    if status=='0':
                        mobilePhone=phone04
                        return(mobilePhone)
                        exit()

import sumeraGetData as SD
s=SD.main()
#a=SD.objList['739135677919730']
#print(a.sumeraThisYearPremia)


#file hendling
outPutRute='/Outputs'
inputFile="BafiExportConverted.xlsx"
outPutFile="fixedBafi.xlsx"

try:
    fileH=open('inputs/'+inputFile)
except:
    print('no such file'+inputFile)
    exit()

#prepering xls's for read and write
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
writeWB = Workbook()
rws = writeWB.active
from openpyxl import load_workbook
readWB = load_workbook('inputs/'+inputFile)
wws = readWB.active

# read the input xls, add valid mobilePhone
excelLine=1
excelColmn=1
mobilePhone=''

for row in wws.values:
    status=row[excelColmn]
    excelColmn=excelColmn+1
    customerName=row[excelColmn]
    excelColmn=excelColmn+1
    customerID=row[excelColmn]
    excelColmn=excelColmn+1
    phones=str(row[excelColmn])
    excelColmn=excelColmn+1
    Address=row[excelColmn]
    excelColmn=excelColmn+1
    email=row[excelColmn]
    excelColmn=excelColmn+1
    policyNumber=row[excelColmn]
    excelColmn=excelColmn+1
    insuranceType=row[excelColmn]
    excelColmn=excelColmn+1
    insuranceCompny=row[excelColmn]
    excelColmn=excelColmn+1
    anafID=row[excelColmn]
    excelColmn=excelColmn+1
    anafName=row[excelColmn]
    excelColmn=excelColmn+1
    policyStartDate=row[excelColmn]
    excelColmn=excelColmn+1
    policyEndDate=row[excelColmn]
    excelColmn=excelColmn+1
    prmia=row[excelColmn]
    excelColmn=excelColmn+1
    veicleType=row[excelColmn]
    excelColmn=excelColmn+1
    SubAnaf=row[excelColmn]
    excelColmn=excelColmn+1
    carLicenessPlate=row[excelColmn]
    excelColmn=excelColmn+1
    manifctureYear=row[excelColmn]
    excelColmn=excelColmn+1
    vehicleModelId=row[excelColmn]
    excelColmn=excelColmn+1
    vehicleModelDescription=row[excelColmn]
    excelColmn=excelColmn+1
    allowedToDrive=row[excelColmn]
    excelColmn=excelColmn+1
    protuction=row[excelColmn]
    excelColmn=excelColmn+1
    propertyAddress=row[excelColmn]
    excelColmn=excelColmn+1
    paymentType=row[excelColmn]
    excelColmn=excelColmn+1
    numofPayments=row[excelColmn]
    excelColmn=excelColmn+1
    egentName=row[excelColmn]
    excelColmn=excelColmn+1
    egentId=row[excelColmn]
    excelColmn=excelColmn+1
    ipdatedBy=row[excelColmn]
    excelColmn=excelColmn+1
    crmRepName=row[excelColmn]
    excelColmn=excelColmn+1
    salesRepName=row[excelColmn]
    excelColmn=excelColmn+1
    comments=row[excelColmn]
    excelColmn=excelColmn+1
    policyhandkerName=row[excelColmn]
    mobilePhone=phoneSplitter(phones) # get a valid mobilePhone


    #writing to the outPutFile and save
    excelColmn=1
    d = rws.cell(row=excelLine, column=excelColmn, value=status)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=customerName)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=customerID)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=phones)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=Address)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=email)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=policyNumber)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=insuranceType)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=insuranceCompny)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=anafID)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=anafName)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=policyStartDate)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=policyEndDate)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=prmia)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=veicleType)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=SubAnaf)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=carLicenessPlate)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=manifctureYear)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=vehicleModelId)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=vehicleModelDescription)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=allowedToDrive)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=protuction)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=vehicleModelDescription)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=propertyAddress)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=paymentType)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=numofPayments)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=egentName)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=egentId)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=ipdatedBy)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=crmRepName)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=salesRepName)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=comments)
    excelColmn=excelColmn+1
    d = rws.cell(row=excelLine, column=excelColmn, value=policyhandkerName)
    excelColmn=excelColmn+1
    if excelLine==1:
        mobilePhone='מספר נייד'
        d = rws.cell(row=excelLine, column=excelColmn, value=mobilePhone)
        excelColmn=excelColmn+1

        d = rws.cell(row=excelLine, column=excelColmn, value='sumeraThisYearPremia')
        excelColmn=excelColmn+1
    excelColmn=excelColmn+1
    else:
        d = rws.cell(row=excelLine, column=excelColmn, value=mobilePhone)
        excelColmn=excelColmn+1

        sdkey=str(policyNumber)+str(anafID) # adding sumera "fileds" if exsist
        spremia=SD.objList
        if sdkey in spremia:
            spremia=SD.objList[sdkey]
            d = rws.cell(row=excelLine, column=excelColmn, value=spremia.sumeraThisYearPremia)
            excelColmn=excelColmn+1
            print('find prmia', spremia.sumeraThisYearPremia)

        else:
            #print('missing key',sdkey)
            a=1+1


    #print(excelLine,mobilePhone,customerName) #this is only to disply something to the user
    excelLine=excelLine+1
    excelColmn=1
writeWB.save('Outputs/'+outPutFile)
